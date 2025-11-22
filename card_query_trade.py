"""
脚本名称：北理工校园卡交易查询
作者：Gemini for User
描述：本地 Python 脚本，用于查询校园卡交易流水并导出为 CSV/Excel。
用法：
1. 确保已安装依赖: pip install -r requirements.txt
2. 配置 .env 或环境变量 (CARD_OPENID, CARD_JSESSIONID)
3. 运行: python card_query_trade.py
"""
import os
import json
import sys
import csv
import re
import argparse
import requests
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from dotenv import load_dotenv

# 加载环境变量
load_dotenv()

def get_cookies_from_gist():
    """尝试从 GitHub Gist 获取最新的 Cookie"""
    github_token = os.getenv("GITHUB_TOKEN")
    gist_id = os.getenv("GIST_ID")
    gist_filename = os.getenv("GIST_FILENAME", "bit_card_cookies.json")

    if not github_token or not gist_id:
        return None, None

    print("正在尝试从 Gist 获取最新凭证...")
    headers = {
        "Authorization": f"token {github_token}",
        "Accept": "application/vnd.github.v3+json"
    }
    
    try:
        # 忽略 SSL 验证警告 (解决本地 SSL 报错问题)
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

        response = requests.get(f"https://api.github.com/gists/{gist_id}", headers=headers, timeout=10, verify=False)
        response.raise_for_status()
        data = response.json()
        
        files = data.get("files", {})
        if gist_filename in files:
            content = files[gist_filename].get("content")
            if content:
                cookie_data = json.loads(content)
                jsessionid = cookie_data.get("jsessionid")
                openid = cookie_data.get("openid")
                print(f"成功从 Gist 获取凭证 (更新时间: {cookie_data.get('updated_at')})")
                return jsessionid, openid
    except Exception as e:
        print(f"从 Gist 获取凭证失败: {e}")
    
    return None, None

def get_current_balance(session, openid, jsessionid):
    """获取当前校园卡余额"""
    url = f"https://dkykt.info.bit.edu.cn/home/openHomePage?openid={openid}"
    headers = {
        "Host": "dkykt.info.bit.edu.cn",
        "User-Agent": "Mozilla/5.0 (iPad; CPU OS 18_5 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Mobile/22F76 AliApp(DingTalk/7.6.46) dd.work.exclusive4bitding/702517158747538432 Channel/exclusive_dingtalk_186038178 Pad/iPad Device/exclusive_dingtalk_186038178 exclusive_dingtalk_186038178/7.6.46 2ndType/exclusive language/zh-Hans-CN UT4Aplus/0.0.6 WK",
        "Cookie": f"JSESSIONID={jsessionid}",
        "Referer": f"https://dkykt.info.bit.edu.cn/selftrade/openQueryCardSelfTrade?openid={openid}&displayflag=1&id=19"
    }
    
    try:
        response = session.get(url, headers=headers, timeout=10)
        response.raise_for_status()
        
        if "openid无效" in response.text or "页面丢失" in response.text:
            print("错误: 页面提示 openid 无效或页面丢失。请检查 openid 是否过期。")
            sys.exit(1)

        # 查找 <p1 id="hidebalanceid" style="display: none">813.42</p1>
        match = re.search(r'<p1 id="hidebalanceid"[^>]*>([\d\.]+)</p1>', response.text)
        if match:
            return float(match.group(1))
        
        # 备用查找 <span name="showbalanceid">余额:￥813.42</span>
        match = re.search(r'<span name="showbalanceid">余额:￥([\d\.]+)</span>', response.text)
        if match:
            return float(match.group(1))
            
        print("警告: 未能从页面解析出余额")
        # print(f"页面内容片段: {response.text[:200]}") # 调试用
        return None
    except Exception as e:
        print(f"获取余额失败: {e}")
        return None

def fetch_trade_data(session, base_url, openid, jsessionid, start_date, end_date):
    """查询指定时间段的流水数据"""
    fmt_date = "%Y-%m-%d"
    str_start_date = start_date.strftime(fmt_date)
    str_end_date = end_date.strftime(fmt_date)
    
    url = f"{base_url}?openid={openid}"
    
    headers = {
        "Host": "dkykt.info.bit.edu.cn",
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "x-requested-with": "XMLHttpRequest",
        "Sec-Fetch-Site": "same-origin",
        "Accept-Language": "zh-CN,zh-Hans;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "Sec-Fetch-Mode": "cors",
        "Content-Type": "application/json",
        "Origin": "https://dkykt.info.bit.edu.cn",
        "User-Agent": "Mozilla/5.0 (iPad; CPU OS 18_5 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Mobile/22F76 AliApp(DingTalk/7.6.46) dd.work.exclusive4bitding/702517158747538432 Channel/exclusive_dingtalk_186038178 Pad/iPad Device/exclusive_dingtalk_186038178 exclusive_dingtalk_186038178/7.6.46 2ndType/exclusive language/zh-Hans-CN UT4Aplus/0.0.6 WK",
        "Referer": f"https://dkykt.info.bit.edu.cn/selftrade/openQueryCardSelfTrade?openid={openid}&displayflag=1&id=19",
        "DingTalk-Flag": "1",
        "Connection": "keep-alive",
        "Cookie": f"JSESSIONID={jsessionid}"
    }

    payload = {
        "beginDate": str_start_date,
        "endDate": str_end_date,
        "tradeType": "-1",
        "openid": openid,
        "idserialOther": "",
        "chooseZH": "1"
    }

    try:
        response = session.post(url, headers=headers, json=payload, timeout=10)
        response.raise_for_status()
        data = response.json()
        if data.get("success"):
            return data.get("resultData", [])
        else:
            print(f"[{str_start_date} ~ {str_end_date}] 查询失败: {data.get('message')}")
            return []
    except Exception as e:
        print(f"[{str_start_date} ~ {str_end_date}] 请求异常: {e}")
        return []

def classify_trade(item):
    """对交易进行分类"""
    tx_amt = float(item.get("txamt", 0))
    tx_name = item.get("txname", "")
    mer_name = item.get("mername", "")
    
    if "补助" in tx_name or "补助" in mer_name:
        return "补助"
    elif tx_amt > 0:
        return "充值"
    else:
        return "消费"

def export_to_excel(trades, filename):
    """导出数据到 Excel 文件，包含多个 Sheet"""
    if not trades:
        print("没有数据可导出")
        return

    try:
        wb = openpyxl.Workbook()
        
        # 定义 Sheet 名称和对应的数据筛选逻辑
        sheets_config = {
            "all": lambda x: True,
            "消费": lambda x: classify_trade(x) == "消费",
            "充值": lambda x: classify_trade(x) == "充值",
            "补助": lambda x: classify_trade(x) == "补助"
        }

        # 移除默认创建的 Sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        headers = ['交易时间', '商户名称', '交易金额', '交易分类', '交易类型', '流水号', '余额']

        for sheet_name, filter_func in sheets_config.items():
            ws = wb.create_sheet(title=sheet_name)
            ws.append(headers)
            
            # 筛选数据
            sheet_data = [item for item in trades if filter_func(item)]
            
            for item in sheet_data:
                category = classify_trade(item)
                # 尝试将金额转换为浮点数，以便 Excel 正确处理数字
                try:
                    amount = float(item.get("txamt", 0))
                except ValueError:
                    amount = item.get("txamt", 0)
                
                # 尝试将余额转换为浮点数
                try:
                    balance = float(item.get("balance_after", 0)) if item.get("balance_after") is not None else ""
                except ValueError:
                    balance = item.get("balance_after", "")

                ws.append([
                    item.get("txdate", ""),
                    item.get("mername", ""),
                    amount,
                    category,
                    item.get("txname", ""),
                    item.get("posjourno") or item.get("journo", ""),
                    balance
                ])
            
            # 调整列宽
            for col_idx, column_cells in enumerate(ws.columns, 1):
                max_length = 0
                column = get_column_letter(col_idx)
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

        wb.save(filename)
        print(f"成功导出数据到 Excel: {filename}")
    except Exception as e:
        print(f"导出 Excel 失败: {e}")

def export_to_csv(trades, filename):
    """导出数据到 CSV 文件"""
    if not trades:
        print("没有数据可导出")
        return

    try:
        # 使用 utf-8-sig 方便 Excel 打开
        with open(filename, mode='w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            # 写入表头
            writer.writerow(['交易时间', '商户名称', '交易金额', '交易分类', '交易类型', '流水号', '余额'])
            
            for item in trades:
                category = classify_trade(item)
                writer.writerow([
                    item.get("txdate", ""),
                    item.get("mername", ""),
                    item.get("txamt", ""),
                    category,
                    item.get("txname", ""),
                    item.get("posjourno") or item.get("journo", ""),
                    item.get("balance_after", "")
                ])
        print(f"成功导出数据到: {filename}")
    except Exception as e:
        print(f"导出失败: {e}")

def main():
    parser = argparse.ArgumentParser(description="北理工校园卡流水查询工具")
    
    parser.add_argument("-s", "--start", help="查询开始日期 (YYYY-MM-DD)")
    parser.add_argument("-e", "--end", help="查询结束日期 (YYYY-MM-DD)，默认为今天")
    parser.add_argument("-d", "--days", type=int, default=60, help="查询最近多少天 (默认 60 天)")
    parser.add_argument("-c", "--category", choices=['消费', '充值', '补助', 'all'], default='all', help="筛选交易类型 (默认 all)")
    parser.add_argument("-b", "--balance", type=float, help="当前校园卡余额 (如果不提供，尝试自动获取)")
    parser.add_argument("-o", "--output", help="导出结果到 CSV 文件路径")
    
    # 如果没有参数，显示帮助
    if len(sys.argv) == 1:
        parser.print_help()
        return

    args = parser.parse_args()

    # 1. 获取配置信息
    # 优先从 Gist 获取
    jsessionid, openid = get_cookies_from_gist()
    
    # 如果 Gist 获取失败，回退到环境变量
    if not jsessionid or not openid:
        print("使用本地环境变量配置...")
        openid = os.getenv("CARD_OPENID")
        jsessionid = os.getenv("CARD_JSESSIONID")

    if not openid or not jsessionid:
        print("错误: 请在 .env 文件或环境变量中设置 CARD_OPENID 和 CARD_JSESSIONID")
        return

    base_url = "https://dkykt.info.bit.edu.cn/selftrade/queryCardSelfTradeList"
    session = requests.Session()

    # 获取当前余额
    current_balance = args.balance
    if current_balance is None:
        current_balance = get_current_balance(session, openid, jsessionid)
    
    if current_balance is not None:
        print(f"当前余额: {current_balance:.2f} 元")
    else:
        print("未能获取当前余额，将无法计算历史余额。")

    # 2. 设置查询时间范围
    end_date = datetime.now()
    if args.end:
        try:
            end_date = datetime.strptime(args.end, "%Y-%m-%d")
            # 如果指定了结束日期，且没有指定时间，默认设为当天的 23:59:59 (虽然接口只看日期，但逻辑上保持一致)
            end_date = end_date.replace(hour=23, minute=59, second=59)
        except ValueError:
            print("错误: 结束日期格式不正确，应为 YYYY-MM-DD")
            return

    if args.start:
        try:
            start_date = datetime.strptime(args.start, "%Y-%m-%d")
        except ValueError:
            print("错误: 开始日期格式不正确，应为 YYYY-MM-DD")
            return
    else:
        # 如果没有指定开始日期，使用 days 参数
        start_date = end_date - timedelta(days=args.days)
    
    print(f"正在查询: {start_date.strftime('%Y-%m-%d')} 至 {end_date.strftime('%Y-%m-%d')} (分段查询)")

    # base_url 和 session 已在前面初始化
    
    all_trades = []
    seen_ids = set() # 用于去重

    # 3. 分段查询 (每7天一段，避免超过100条限制)
    current_start = start_date
    step_days = 7
    
    while current_start < end_date:
        current_end = min(current_start + timedelta(days=step_days), end_date)
        
        # print(f"查询片段: {current_start.strftime('%Y-%m-%d')} ~ {current_end.strftime('%Y-%m-%d')}")
        trades = fetch_trade_data(session, base_url, openid, jsessionid, current_start, current_end)
        
        if len(trades) >= 100:
            print(f"警告: {current_start.strftime('%Y-%m-%d')} ~ {current_end.strftime('%Y-%m-%d')} 期间记录达到100条上限，可能存在遗漏！")

        for item in trades:
            # 使用 posjourno (POS流水号) 或 journo (流水号) 作为唯一标识
            unique_id = item.get("posjourno") or item.get("journo")
            if unique_id and unique_id not in seen_ids:
                seen_ids.add(unique_id)
                all_trades.append(item)
        
        current_start = current_end # 下一段开始

    # 4. 结果处理与展示
    # 按时间倒序排序
    all_trades.sort(key=lambda x: x.get("txdate", ""), reverse=True)

    # 计算历史余额
    if current_balance is not None:
        temp_balance = current_balance
        for item in all_trades:
            item["balance_after"] = f"{temp_balance:.2f}"
            try:
                amt = float(item.get("txamt", 0))
                temp_balance -= amt
            except:
                pass
    
    # 筛选
    filtered_trades = []
    if args.category == 'all':
        filtered_trades = all_trades
    else:
        for item in all_trades:
            if classify_trade(item) == args.category:
                filtered_trades.append(item)

    print(f"\n查询完成，共获取 {len(filtered_trades)} 条记录 (总记录 {len(all_trades)} 条)：")
    print("-" * 90)
    print(f"{'时间':<20} | {'商户':<15} | {'金额':<8} | {'余额':<8} | {'分类':<6} | {'类型'}")
    print("-" * 90)
    
    stats = {"消费": 0.0, "充值": 0.0, "补助": 0.0}
    
    for item in filtered_trades:
        tx_date = item.get("txdate", "")
        mer_name = item.get("mername", "")
        tx_amt = item.get("txamt", "0")
        tx_name = item.get("txname", "")
        balance_str = item.get("balance_after", "-")
        
        category = classify_trade(item)
        # 即使筛选了，统计时只统计筛选后的
        stats[category] += float(tx_amt)
        
        print(f"{tx_date:<20} | {mer_name:<15} | {tx_amt:<8} | {balance_str:<8} | {category:<6} | {tx_name}")

    print("-" * 90)
    print("当前筛选结果统计:")
    if args.category == 'all' or args.category == '消费':
        print(f"总消费: {stats['消费']:.2f} 元")
    if args.category == 'all' or args.category == '充值':
        print(f"总充值: {stats['充值']:.2f} 元")
    if args.category == 'all' or args.category == '补助':
        print(f"总补助: {stats['补助']:.2f} 元")
    
    if args.category == 'all':
        print(f"净收支: {sum(stats.values()):.2f} 元")

    # 5. 导出
    if args.output:
        if args.output.endswith('.xlsx'):
            export_to_excel(all_trades, args.output)
        else:
            export_to_csv(filtered_trades, args.output)

if __name__ == "__main__":
    main()
